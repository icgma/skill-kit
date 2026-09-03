#!/usr/bin/env python3
"""
xlsx_tools.py — Formula recalculation and verification for Excel files.

Usage:
  python xlsx_tools.py recalc <file.xlsx>     Recalculate all formulas
  python xlsx_tools.py verify <file.xlsx>      Verify no formula errors remain
"""

import argparse
import glob
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile

# UTF-8 stdio guard — win32 pipes default to the ANSI codepage (cp1252/cp936)
# and CJK sheet names / file paths in JSON output would crash the print.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# Forbidden functions (incompatible with Excel 2019 and earlier)
FORBIDDEN_FUNCTIONS = [
    'FILTER', 'UNIQUE', 'SORT', 'SORTBY', 'XLOOKUP', 'XMATCH',
    'SEQUENCE', 'LET', 'LAMBDA', 'RANDARRAY'
]

# Implicit array formula pattern
IMPLICIT_ARRAY_PATTERN = re.compile(r'MATCH\s*\(\s*TRUE\s*\(\s*\)', re.IGNORECASE)


def _find_libreoffice():
    """Return a LibreOffice/soffice executable path, or None.

    Discovery order: PATH (`libreoffice`, then `soffice`), then platform
    app paths (macOS .app bundle, Windows Program Files, common Linux paths).
    """
    candidates = []
    for name in ('libreoffice', 'soffice'):
        found = shutil.which(name)
        if found:
            candidates.append(found)
    if sys.platform == 'darwin':
        candidates.append('/Applications/LibreOffice.app/Contents/MacOS/soffice')
    elif sys.platform == 'win32':
        for env_var in ('PROGRAMFILES', 'PROGRAMFILES(X86)', 'LOCALAPPDATA'):
            base = os.environ.get(env_var)
            if base:
                # soffice.com is the console-attached variant: it stays
                # synchronous and keeps stdout/stderr attached, whereas
                # GUI-subsystem soffice.exe may return before --convert-to
                # finishes. Prefer .com; .exe is the fallback.
                candidates.append(os.path.join(base, 'LibreOffice', 'program', 'soffice.com'))
                candidates.append(os.path.join(base, 'LibreOffice', 'program', 'soffice.exe'))
    else:  # Linux and other POSIX
        candidates.extend(['/usr/bin/soffice', '/usr/local/bin/soffice',
                           '/snap/bin/libreoffice', '/opt/libreoffice/program/soffice'])
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None


def _conditional_formatting_count(filepath):
    """Count conditional-formatting rule blocks in sheet XML (std + x14 ext)."""
    total = 0
    try:
        import zipfile
        with zipfile.ZipFile(filepath) as z:
            for name in z.namelist():
                if re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                    xml = z.read(name).decode('utf-8', errors='replace')
                    total += xml.count('<conditionalFormatting')
                    total += xml.count(':conditionalFormatting')  # x14 extension blocks
    except Exception:
        pass
    return total


# LibreOffice headless does NOT recalculate xlsx cells that already carry a
# cached <v> value: the OOXMLRecalcMode load option defaults to 'never', so
# stale caches written by Excel/WPS used to pass through unchanged while we
# reported recalculated=true. Pre-seeding the isolated profile with
# OOXMLRecalcMode=0 forces a full recalculation on load.
_LO_FORCE_RECALC_XCU = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
    'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
    'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
    ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>'
    '</item>\n</oor:items>\n'
)


def _seed_lo_profile(profile_dir):
    """Pre-seed an isolated LO user profile forcing full recalc on xlsx load.
    Failure is non-fatal: unseeded only means stale caches may survive."""
    try:
        user_dir = os.path.join(profile_dir, 'user')
        os.makedirs(user_dir, exist_ok=True)
        with open(os.path.join(user_dir, 'registrymodifications.xcu'),
                  'w', encoding='utf-8') as f:
            f.write(_LO_FORCE_RECALC_XCU)
    except OSError:
        pass


def cmd_recalc(args):
    """Recalculate all formulas in an Excel file."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(json.dumps({"status": "error", "message": f"File not found: {filepath}"}))
        sys.exit(2)

    # Hard gate: .xlsx only. Without this, `recalc data.csv` makes
    # LibreOffice convert the CSV to xlsx and then overwrite the ORIGINAL
    # .csv path with binary content -- silent data destruction reported as
    # success. (.xlsm is refused for the same reason: conversion strips VBA.)
    if os.path.splitext(filepath)[1].lower() != '.xlsx':
        print(json.dumps({
            "status": "error",
            "message": (f"recalc only supports .xlsx files, got: {filepath}. "
                        "Convert to .xlsx first (e.g. open in Excel/WPS and Save As .xlsx)."),
        }))
        sys.exit(2)

    # Option 1: formulas library
    # NOTE: the `formulas` package is NOT installed in the managed runtime and
    # this branch is UNTESTED — in particular xl_model.write(dirpath=...) may
    # expect a directory rather than a file path. Kept as a best-effort path;
    # do not rely on it without installing `formulas` and re-testing.
    try:
        import formulas
        print("Recalculating with formulas library...")
        xl_model = formulas.ExcelModel().loads(filepath).finish()
        xl_model.calculate()

        # Write to temp file for verification only (formulas strips charts/styles)
        base, ext = os.path.splitext(filepath)
        verify_path = (base + '_recalc_verify.xlsx') if ext.lower() == '.xlsx' else (filepath + '_recalc_verify.xlsx')
        xl_model.write(dirpath=verify_path)
        print(json.dumps({
            "status": "success",
            "method": "formulas",
            "recalculated": True,
            "delivery_file_recalculated": False,
            "verification_file": verify_path,
            "note": f"Verification file: {verify_path} (do NOT deliver this — it strips charts/styles). Deliver the original file.",
        }))

        # Set calcMode=auto on the ORIGINAL file so Excel recalculates on open
        _set_calc_mode_auto(filepath)
        return
    except ImportError:
        print("formulas library not available, trying libreoffice...")
    except Exception as e:
        print(f"formulas library failed: {e}, trying libreoffice...")

    # Option 2: libreoffice headless.
    # LibreOffice refuses to overwrite the source file when converting to the
    # same format in the same directory, so convert into a fresh temp dir and
    # move the result back over the original.
    lo_bin = _find_libreoffice()
    if lo_bin:
        cf_before = _conditional_formatting_count(filepath)
        tmpdir = tempfile.mkdtemp(prefix='xlsx_recalc_')
        profile = tempfile.mkdtemp(prefix='xlsx_lo_profile_')
        _seed_lo_profile(profile)
        try:
            from pathlib import Path
            subprocess.run([
                lo_bin, '--headless', '--calc',
                # Isolated user profile: a locked/absent default profile
                # (fresh CI machines, parallel runs) must not break recalc.
                f'-env:UserInstallation={Path(profile).resolve().as_uri()}',
                '--convert-to', 'xlsx',
                '--outdir', tmpdir, os.path.abspath(filepath)
            ], check=True, capture_output=True, timeout=120)
            converted = glob.glob(os.path.join(tmpdir, '*.xlsx'))
            if not converted:
                raise RuntimeError('LibreOffice produced no output file')
            # os.replace cannot move across volumes — on win32 the temp dir
            # (C:\Users\...\Temp) and the checkout/target (D:\...) are often
            # different drives and it dies with WinError 17. Copy bytes;
            # the finally block removes the temp dir.
            shutil.copyfile(converted[0], filepath)
            warnings = []
            if cf_before > 0 and _conditional_formatting_count(filepath) < cf_before:
                warnings.append('LibreOffice recalc may drop conditional-formatting extensions')
            print(json.dumps({
                "status": "success",
                "method": "libreoffice",
                "recalculated": True,
                "delivery_file_recalculated": True,
                "warnings": warnings,
            }))
            return
        except Exception as e:
            print(f"libreoffice failed: {e}, falling back to calcMode=auto...")
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)
            shutil.rmtree(profile, ignore_errors=True)

    # Fallback: request recalculation on open, but do not claim that formulas ran.
    if not _set_calc_mode_auto(filepath):
        print(json.dumps({
            "status": "error",
            "method": "calcMode_auto",
            "recalculated": False,
            "message": "No calculation engine is available and calcMode=auto could not be set.",
        }))
        sys.exit(2)
    print(json.dumps({
        "status": "deferred",
        "method": "calcMode_auto",
        "recalculated": False,
        "delivery_file_recalculated": False,
        "note": "Neither formulas nor libreoffice is available. Formula results remain unverified; Excel/WPS has been asked to recalculate them on open.",
    }))


def _set_calc_mode_auto(filepath):
    """Set calcMode to auto so Excel recalculates on open."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        if wb.calculation is None:
            from openpyxl.workbook.properties import CalcProperties
            wb.calculation = CalcProperties()
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Warning: could not set calcMode: {e}", file=sys.stderr)
        return False


def cmd_verify(args):
    """Verify an Excel file for formula errors, forbidden functions, and implicit arrays."""
    filepath = args.file
    if not os.path.exists(filepath):
        print(json.dumps({"status": "error", "message": f"File not found: {filepath}"}))
        sys.exit(2)

    from openpyxl import load_workbook

    issues = []
    wb_data = None
    formula_count = 0
    missing_formula_cache_locations = []

    # Check 1: Formula errors (open with data_only to see cached values)
    try:
        wb_data = load_workbook(filepath, data_only=True)
        for ws in wb_data.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('#'):
                        issues.append({
                            "type": "formula_error",
                            "severity": "error",
                            "location": f"{ws.title}!{cell.coordinate}",
                            "value": cell.value,
                        })
    except Exception as e:
        issues.append({"type": "load_error", "severity": "error", "detail": str(e)})

    # Check 2: Forbidden functions + implicit arrays (open with formulas)
    try:
        wb_formulas = load_workbook(filepath)
        for ws in wb_formulas.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        formula = cell.value.upper()

                        if (
                            wb_data is not None
                            and ws.title in wb_data.sheetnames
                            and wb_data[ws.title][cell.coordinate].value is None
                        ):
                            missing_formula_cache_locations.append(
                                f"{ws.title}!{cell.coordinate}"
                            )

                        # Forbidden functions
                        for func in FORBIDDEN_FUNCTIONS:
                            if func + '(' in formula:
                                issues.append({
                                    "type": "forbidden_function",
                                    "severity": "warning",
                                    "location": f"{ws.title}!{cell.coordinate}",
                                    "function": func,
                                    "formula": cell.value[:80],
                                })

                        # Implicit array formula
                        if IMPLICIT_ARRAY_PATTERN.search(cell.value):
                            issues.append({
                                "type": "implicit_array",
                                "severity": "warning",
                                "location": f"{ws.title}!{cell.coordinate}",
                                "formula": cell.value[:80],
                                "hint": "Use SUMPRODUCT or helper column instead of MATCH(TRUE(), ...)",
                            })
    except Exception as e:
        issues.append({"type": "load_error", "severity": "error", "detail": str(e)})

    if missing_formula_cache_locations:
        sample_limit = 20
        issues.append({
            "type": "formula_cache_missing",
            "severity": "warning",
            "count": len(missing_formula_cache_locations),
            "locations": missing_formula_cache_locations[:sample_limit],
            "omitted_location_count": max(
                0, len(missing_formula_cache_locations) - sample_limit
            ),
            "detail": "Formula cells have no cached results, so their outcomes were not evaluated.",
            "hint": "Recalculate with Excel/WPS/LibreOffice or independently verify every reported derived value from the raw inputs.",
        })

    # Output
    errors = [i for i in issues if i['severity'] == 'error']
    warnings = [i for i in issues if i['severity'] == 'warning']

    has_missing_formula_cache = any(
        issue.get("type") == "formula_cache_missing" for issue in issues
    )
    status = (
        "fail"
        if errors
        else "unverified"
        if has_missing_formula_cache
        else "pass_with_warnings"
        if warnings
        else "pass"
    )

    result = {
        "status": status,
        "file": filepath,
        "formula_count": formula_count,
        "formula_cache_missing_count": len(missing_formula_cache_locations),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "issues": issues,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if not errors else 1)


def main():
    parser = argparse.ArgumentParser(description="Excel formula tools")
    subparsers = parser.add_subparsers(dest="command")

    recalc_p = subparsers.add_parser("recalc", help="Recalculate all formulas")
    recalc_p.add_argument("file", help="Excel file path")
    recalc_p.set_defaults(func=cmd_recalc)

    verify_p = subparsers.add_parser("verify", help="Verify formula errors and forbidden functions")
    verify_p.add_argument("file", help="Excel file path")
    verify_p.set_defaults(func=cmd_verify)

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(0)
    args.func(args)


if __name__ == "__main__":
    main()
